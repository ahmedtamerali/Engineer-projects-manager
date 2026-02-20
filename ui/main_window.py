import os
import shutil
from datetime import datetime
import ttkbootstrap as tb
from ttkbootstrap.constants import RIGHT, LEFT
from tkinter import messagebox
from tkinter import simpledialog
from ui.project_window import ProjectWindow, WorkerDetailWindow
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class MainWindow:
    def __init__(self, db):
        self.db = db
        self.app = tb.Window(themename='darkly')
        self.app.title('projects manager - Eng/Abdelarahman')
        self.app.geometry('900x640')
        self.app.resizable(False, False)
        self.app.option_add('*Font', 'SegoeUI 10')
        self._build_ui()

    def _build_ui(self):
        top = tb.Frame(self.app)
        top.pack(fill='x', padx=12, pady=(8, 0))

        title = tb.Label(top, text='مدير المشاريع والمدفوعات', font=('Segoe UI', 16, 'bold'), anchor='e')
        title.pack(side='right', fill='both', expand=True, padx=(0, 40))

        actions = tb.Frame(top)
        actions.pack(side='right')

        export_btn = tb.Button(actions, text='تصدير إلى Excel', bootstyle='info-outline', command=self.export_to_excel)
        export_btn.pack(side=LEFT, padx=6)

        backup_btn = tb.Button(actions, text='نسخ احتياطي لقاعدة البيانات', bootstyle='secondary-outline', command=self.backup_db)
        backup_btn.pack(side=LEFT, padx=6)

        add_btn = tb.Button(self.app, text='+', bootstyle='success', width=3, command=self.add_project_dialog)
        add_btn.place(relx=0.95, rely=0.02)

        # Scrollable area for cards
        container = tb.Frame(self.app)
        container.pack(fill='both', expand=True, padx=12, pady=(8, 12))

        self.canvas = tb.Canvas(container, bg=self.app.cget('bg'), highlightthickness=0)
        scrollbar = tb.Scrollbar(container, orient='vertical', command=self.canvas.yview)
        scrollbar.pack(side='right', fill='y')
        self.canvas.pack(side='left', fill='both', expand=True)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        self.cards_frame = tb.Frame(self.canvas)
        canvas_window = self.canvas.create_window((0, 0), window=self.cards_frame, anchor='nw')

        def resize_canvas(event):
            self.canvas.itemconfig(canvas_window, width=event.width)

        self.cards_frame.bind('<Configure>', lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))
        self.canvas.bind('<Configure>', resize_canvas)
        
        # Configure grid columns to expand equally
        self.cards_frame.columnconfigure(0, weight=1)
        self.cards_frame.columnconfigure(1, weight=1)

        self.cards_frame.bind('<Configure>', lambda e: self.canvas.configure(scrollregion=self.canvas.bbox('all')))
        
        # Add mouse wheel scrolling support
        self._setup_canvas_scrolling(self.canvas, self.cards_frame)

        self.load_projects()

    def _setup_canvas_scrolling(self, canvas, frame):
        """Enable mouse wheel scrolling on the canvas and all child frames"""
        def _on_mousewheel(event):
            # Ensure canvas has focus for smooth scrolling
            canvas.focus_set()
            
            # Handle Windows and Linux mouse wheel events with smooth scrolling
            try:
                if event.num == 5 or event.delta < 0:
                    canvas.yview_scroll(3, "units")
                elif event.num == 4 or event.delta > 0:
                    canvas.yview_scroll(-3, "units")
            except Exception:
                pass
        
        # Bind to canvas
        canvas.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Button-4>", _on_mousewheel)
        canvas.bind("<Button-5>", _on_mousewheel)
        
        # Bind to main frame
        frame.bind("<MouseWheel>", _on_mousewheel)
        frame.bind("<Button-4>", _on_mousewheel)
        frame.bind("<Button-5>", _on_mousewheel)
        
        # Recursively bind to all child frames
        def bind_children(parent):
            for child in parent.winfo_children():
                if isinstance(child, tb.Frame):
                    child.bind("<MouseWheel>", _on_mousewheel)
                    child.bind("<Button-4>", _on_mousewheel)
                    child.bind("<Button-5>", _on_mousewheel)
                    bind_children(child)
        
        bind_children(frame)

    def load_projects(self):
        for w in self.cards_frame.winfo_children():
            w.destroy()

        projects = self.db.get_all_projects()
        cols = 2
        r = c = 0
        
        # Load projects first
        for p in projects:
            frame = tb.Frame(self.cards_frame, padding=12, relief='raised', bootstyle='secondary')
            frame.grid(row=r, column=c, padx=8, pady=8, sticky='nsew')

            name = tb.Label(frame, text=p['name'], font=('Segoe UI', 12, 'bold'), anchor='e')
            name.pack(fill='x')

            # Get workers + importers totals only (excluding customer)
            try:
                total, paid = self.db.get_workers_importers_summary(p['id'])
            except Exception:
                total = 0
                paid = 0
            remain = total - paid

            lbl_total = tb.Label(frame, text=f'المبلغ الكلي المُكلّف: {total:.2f}', anchor='e')
            lbl_total.pack(fill='x')
            lbl_paid = tb.Label(frame, text=f'المبلغ المدفوع: {paid:.2f}', anchor='e')
            lbl_paid.pack(fill='x')
            lbl_remain = tb.Label(frame, text=f'المبلغ المتبقي: {remain:.2f}', anchor='e')
            lbl_remain.pack(fill='x')

            def on_left(e, pid=p['id'], pname=p['name']):
                ProjectWindow(self.db, pid, self.load_projects, pname)

            def on_right(e, pid=p['id'], title=p['name']):
                menu = tb.Menu(self.app, tearoff=0)
                menu.add_command(label='تعديل اسم المشروع', command=lambda: self.edit_project_dialog(pid))
                menu.add_command(label='حذف المشروع', command=lambda: self.delete_project(pid))
                menu.tk_popup(e.x_root, e.y_root)

            frame.bind('<Button-1>', on_left)
            frame.bind('<Button-3>', on_right)

            c += 1
            if c >= cols:
                c = 0
                r += 1
        
        # Add separator / section header for workers and importers
        if projects:  # Only add separator if there are projects
            r += 1
        
        header_frame = tb.Frame(self.cards_frame)
        header_frame.grid(row=r, column=0, columnspan=cols, sticky='ew', padx=8, pady=(12, 6))
        
        section_label = tb.Label(header_frame, text='العاملون والموردون', 
                                font=('Segoe UI', 12, 'bold'), anchor='e')
        section_label.pack(side='right', fill='x', expand=True)
        
        r += 1
        c = 0
        
        # Load workers
        try:
            workers = self.db.get_all_workers_with_totals()
            for w in workers:
                frame = tb.Frame(self.cards_frame, padding=12, relief='raised', bootstyle='info')
                frame.grid(row=r, column=c, padx=8, pady=8, sticky='nsew')

                # Name with job type
                name_text = f"{w['name']}"
                if w['job']:
                    name_text += f" ({w['job']})"
                name = tb.Label(frame, text=name_text, font=('Segoe UI', 11, 'bold'), anchor='e')
                name.pack(fill='x')
                
                # Type label
                type_label = tb.Label(frame, text='عامل', font=('Segoe UI', 9), anchor='e', foreground='gray')
                type_label.pack(fill='x')

                # Projects list
                if w['projects']:
                    projects_text = ', '.join([p['name'] for p in w['projects']])
                    proj_label = tb.Label(frame, text=f'المشاريع: {projects_text}', font=('Segoe UI', 8), anchor='e', foreground='gray')
                    proj_label.pack(fill='x')

                lbl_total = tb.Label(frame, text=f'المبلغ الكلي المُكلّف: {w["total_assigned"]:.2f}', anchor='e')
                lbl_total.pack(fill='x')
                lbl_paid = tb.Label(frame, text=f'المبلغ المدفوع: {w["total_paid"]:.2f}', anchor='e')
                lbl_paid.pack(fill='x')
                lbl_remain = tb.Label(frame, text=f'المبلغ المتبقي: {w["total_remaining"]:.2f}', anchor='e')
                lbl_remain.pack(fill='x')

                def on_worker_click(e, worker_ids=w['worker_ids'], wname=w['name']):
                    WorkerDetailWindow(self.db, 'worker', worker_ids, wname, self.load_projects)

                frame.bind('<Button-1>', on_worker_click)

                c += 1
                if c >= cols:
                    c = 0
                    r += 1
        except Exception:
            pass
        
        # Load importers
        try:
            importers = self.db.get_all_importers_with_totals()
            for imp in importers:
                frame = tb.Frame(self.cards_frame, padding=12, relief='raised', bootstyle='warning')
                frame.grid(row=r, column=c, padx=8, pady=8, sticky='nsew')

                # Name with job type
                name_text = f"{imp['name']}"
                if imp['job']:
                    name_text += f" ({imp['job']})"
                name = tb.Label(frame, text=name_text, font=('Segoe UI', 11, 'bold'), anchor='e')
                name.pack(fill='x')
                
                # Type label
                type_label = tb.Label(frame, text='مورد', font=('Segoe UI', 9), anchor='e', foreground='gray')
                type_label.pack(fill='x')

                # Projects list
                if imp['projects']:
                    projects_text = ', '.join([p['name'] for p in imp['projects']])
                    proj_label = tb.Label(frame, text=f'المشاريع: {projects_text}', font=('Segoe UI', 8), anchor='e', foreground='gray')
                    proj_label.pack(fill='x')

                lbl_total = tb.Label(frame, text=f'المبلغ الكلي المُكلّف: {imp["total_assigned"]:.2f}', anchor='e')
                lbl_total.pack(fill='x')
                lbl_paid = tb.Label(frame, text=f'المبلغ المدفوع: {imp["total_paid"]:.2f}', anchor='e')
                lbl_paid.pack(fill='x')
                lbl_remain = tb.Label(frame, text=f'المبلغ المتبقي: {imp["total_remaining"]:.2f}', anchor='e')
                lbl_remain.pack(fill='x')

                def on_importer_click(e, importer_ids=imp['importer_ids'], iname=imp['name']):
                    WorkerDetailWindow(self.db, 'importer', importer_ids, iname, self.load_projects)

                frame.bind('<Button-1>', on_importer_click)

                c += 1
                if c >= cols:
                    c = 0
                    r += 1
        except Exception:
            pass
        
        # Configure all rows to expand equally and fill available space
        for i in range(r + 1):
            self.cards_frame.rowconfigure(i, weight=1)
        
        # Rebind scrolling events to newly created frames for smooth scrolling
        def _on_mousewheel(event):
            self.canvas.focus_set()
            try:
                if event.num == 5 or event.delta < 0:
                    self.canvas.yview_scroll(3, "units")
                elif event.num == 4 or event.delta > 0:
                    self.canvas.yview_scroll(-3, "units")
            except Exception:
                pass
        
        def bind_children(parent):
            for child in parent.winfo_children():
                if isinstance(child, tb.Frame):
                    child.bind("<MouseWheel>", _on_mousewheel)
                    child.bind("<Button-4>", _on_mousewheel)
                    child.bind("<Button-5>", _on_mousewheel)
                    bind_children(child)
        
        bind_children(self.cards_frame)

    def add_project_dialog(self):
        try:
            name = simpledialog.askstring('إضافة مشروع', 'اسم المشروع:')
            if name:
                # Check if project name already exists
                projects = self.db.get_all_projects()
                if any(p['name'] == name for p in projects):
                    messagebox.showerror('خطأ', 'اسم المشروع موجود بالفعل!')
                    return
                self.db.add_project(name)
                self.load_projects()
        except Exception:
            messagebox.showerror('خطأ', 'حدث خطأ أثناء الاتصال بقاعدة البيانات.')

    def edit_project_dialog(self, project_id):
        try:
            # fetch current
            projects = self.db.get_all_projects()
            p = next((x for x in projects if x['id'] == project_id), None)
            if not p:
                return
            name = simpledialog.askstring('تعديل المشروع', 'اسم المشروع:', initialvalue=p['name'])
            if name:
                self.db.edit_project(project_id, name)
                self.load_projects()
        except Exception:
            messagebox.showerror('خطأ', 'حدث خطأ أثناء الاتصال بقاعدة البيانات.')

    def delete_project(self, project_id):
        if messagebox.askyesno('تأكيد', 'هل تريد حذف المشروع؟'):
            try:
                self.db.delete_project(project_id)
                self.load_projects()
            except Exception:
                messagebox.showerror('خطأ', 'حدث خطأ أثناء الاتصال بقاعدة البيانات.')

    def export_to_excel(self):
        try:
            projects = self.db.get_all_projects()
            if not projects:
                messagebox.showinfo('تنبيه', 'لا توجد مشاريع للتصدير')
                return
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create summary sheet
            ws_summary = wb.create_sheet('ملخص المشاريع', 0)
            
            # Header styling
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            # Summary headers (reversed for RTL)
            headers = ['اسم المشروع', 'المبلغ المكلف (عمال+موردين)', 'المبلغ المدفوع', 'المبلغ المتبقي']
            cols = ['D', 'C', 'B', 'A']
            for col_letter, header_text in zip(cols, headers):
                cell = ws_summary[f'{col_letter}1']
                cell.value = header_text
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Add summary data (reversed for RTL)
            row = 2
            for p in projects:
                try:
                    total, paid = self.db.get_workers_importers_summary(p['id'])
                except Exception:
                    total, paid = 0, 0
                remain = total - paid
                
                ws_summary[f'D{row}'] = p['name']
                ws_summary[f'C{row}'] = total
                ws_summary[f'B{row}'] = paid
                ws_summary[f'A{row}'] = remain
                row += 1
            
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 20
            ws_summary.column_dimensions['D'].width = 20
            
            # Create detailed sheet for each project
            for p in projects:
                project_id = p['id']
                project_name = p['name'][:30] if len(p['name']) <= 30 else p['name'][:27] + '...'
                
                ws = wb.create_sheet(project_name)
                
                # Project title
                ws['A1'] = f'المشروع: {p["name"]}'
                ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
                ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
                ws.merge_cells('A1:D1')
                
                current_row = 3
                
                # ===== WORKERS SECTION =====
                ws[f'A{current_row}'] = 'العمال'
                ws[f'A{current_row}'].font = Font(bold=True, size=12, color='FFFFFF')
                ws[f'A{current_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                workers = self.db.get_workers_by_project(project_id)
                if workers:
                    # Headers (reversed for RTL)
                    ws[f'D{current_row}'] = 'اسم العامل'
                    ws[f'C{current_row}'] = 'المهنة'
                    ws[f'B{current_row}'] = 'المبلغ المكلف'
                    ws[f'A{current_row}'] = 'المبلغ المدفوع'
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{current_row}'].font = Font(bold=True, color='FFFFFF')
                        ws[f'{col}{current_row}'].fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')
                    current_row += 1
                    
                    for worker in workers:
                        assigns = self.db.get_assignments('worker', worker['id'])
                        total_assigned = sum(a['amount'] for a in assigns)
                        
                        total_paid = 0
                        for a in assigns:
                            pays = self.db.get_payments(a['id'])
                            total_paid += sum(p['amount'] for p in pays)
                        
                        ws[f'D{current_row}'] = worker['name']
                        ws[f'C{current_row}'] = worker.get('job') or ''
                        ws[f'B{current_row}'] = total_assigned
                        ws[f'A{current_row}'] = total_paid
                        current_row += 1
                else:
                    ws[f'A{current_row}'] = 'لا يوجد عمال'
                    current_row += 1
                
                current_row += 1
                
                # ===== IMPORTERS SECTION =====
                ws[f'A{current_row}'] = 'الموردون'
                ws[f'A{current_row}'].font = Font(bold=True, size=12, color='FFFFFF')
                ws[f'A{current_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
                
                importers = self.db.get_importers_by_project(project_id)
                if importers:
                    # Headers (reversed for RTL)
                    ws[f'D{current_row}'] = 'اسم المورد'
                    ws[f'C{current_row}'] = 'السلعة/الوظيفة'
                    ws[f'B{current_row}'] = 'المبلغ المكلف'
                    ws[f'A{current_row}'] = 'المبلغ المدفوع'
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{current_row}'].font = Font(bold=True, color='FFFFFF')
                        ws[f'{col}{current_row}'].fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                    current_row += 1
                    
                    for importer in importers:
                        assigns = self.db.get_assignments('importer', importer['id'])
                        total_assigned = sum(a['amount'] for a in assigns)
                        
                        total_paid = 0
                        for a in assigns:
                            pays = self.db.get_payments(a['id'])
                            total_paid += sum(p['amount'] for p in pays)
                        
                        ws[f'D{current_row}'] = importer['name']
                        ws[f'C{current_row}'] = importer.get('job') or ''
                        ws[f'B{current_row}'] = total_assigned
                        ws[f'A{current_row}'] = total_paid
                        current_row += 1
                else:
                    ws[f'A{current_row}'] = 'لا يوجد موردون'
                    current_row += 1
                
                current_row += 1
                
                # ===== CUSTOMER SECTION =====
                ws[f'A{current_row}'] = 'العميل'
                ws[f'A{current_row}'].font = Font(bold=True, size=12, color='FFFFFF')
                ws[f'A{current_row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                ws.merge_cells(f'A{current_row}:C{current_row}')
                current_row += 1
                
                try:
                    cust_total, cust_paid = self.db.get_customer_summary(project_id)
                except Exception:
                    cust_total, cust_paid = 0, 0
                cust_remain = cust_total - cust_paid
                
                ws[f'D{current_row}'] = 'إجمالي المبلغ المكلف'
                ws[f'C{current_row}'] = cust_total
                current_row += 1
                ws[f'D{current_row}'] = 'المبلغ المدفوع'
                ws[f'C{current_row}'] = cust_paid
                current_row += 1
                ws[f'D{current_row}'] = 'المبلغ المتبقي'
                ws[f'C{current_row}'] = cust_remain
                current_row += 1
                
                # Set column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
            
            # Save file
            fname = f'projects_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(fname)
            messagebox.showinfo('تم', f'تم التصدير إلى {fname}')
        except Exception as e:
            messagebox.showerror('خطأ', f'حدث خطأ أثناء التصدير: {str(e)}')

    def backup_db(self):
        try:
            src = self.db.path
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            dst = f'data_backup_{ts}.db'
            shutil.copyfile(src, dst)
            messagebox.showinfo('تم', f'تم إنشاء النسخة: {dst}')
        except Exception:
            messagebox.showerror('خطأ', 'حدث خطأ أثناء إنشاء النسخة الاحتياطية.')

    def run(self):
        self.app.mainloop()
